# Import stuff that we need
import time
import os
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
import csv

# -----------------------------------------------------BEGIN------------------------------------------------------------
#
# This prevents the browser window from closing automatically. We only need to keep this while we're working on the
# code so that we can observe that things are happening the way we want them to. Delete after program is finished!
# Finished
chrome_options = Options()
chrome_options.add_experimental_option("detach", True)
# ----------------------------------------------------------------------------------------------------------------------
# options=chrome_options
# Get logged in - FINISHED
# Notes:
#   1) driver.get() will eventually use a variable, so that we can use this program with different reports
driver = webdriver.Chrome()
driver.get("https://app.schoox.com/academies/panel/dashboard2/training/course.php?acadId=7592&course_id=3847508")
user = "102026939"
passw = "RedLobster1"
time.sleep(1)
driver.find_element("xpath", "//*[@id='main']/div/main/input[6]").send_keys(user)
driver.find_element("xpath", "//*[@id='main']/div/main/input[7]").send_keys(passw)
button = driver.find_element("name", "button")
button.click()
# # --------------------------------------------------------------------------------------------------------------------
# #
# # # Start pressing buttons to download report - FINISHED
# # Notes:
# #   1) Make sure that xpath is the same on all the different pages. Otherwise, we'll have to figure out a variable.
time.sleep(2)
button = driver.find_element("xpath", "/html/body/div[5]/div/div[2]/div/div[2]/div[2]/div[2]/div/img")
button.click()
button = driver.find_element("xpath", "/html/body/div[5]/div/div[2]/div/div[2]/div[2]/div[2]/div/div/p[1]/a/span/span")
button.click()
time.sleep(4)
button = driver.find_element("xpath", "/html/body/div[7]/div[1]/div[2]/div[2]/div[1]/generate-report/div/div/p/a")
button.click()
time.sleep(1)
# ----------------------------------------------------------------------------------------------------------------------
#
# Select most recently downloaded file, so we can work with it - FINISHED
directory_path = "C:/Users/danny/Downloads/"
most_recent_file = None
most_recent_time = 0

for entry in os.scandir(directory_path):
    if entry.is_file():
        mod_time = entry.stat().st_mtime_ns
        if mod_time > most_recent_time:
            most_recent_file = entry.name
            most_recent_time = mod_time
#
file = f"{directory_path}{most_recent_file}"
# file2 = f"{directory_path}original.csv"
# print(file)
# print(file2)
# ----------------------------------------------------------------------------------------------------------------------
#
# Here, we iterate through each row and save only the ones we want to a new .xlsx file - FINISHED
# Notes:
#   1) Find a way to make unique file names for multiple reports.
#   2) Save new path and filename as a variable to continue working with it
#        - maybe path to new folder, so things don't get all mixed up
wb = openpyxl.Workbook()
ws = wb.active

with open(file) as f:
    reader = csv.reader(f)
    for row in reader:
        try:
            if row[6] == "0" or row[6] == "% Completed":
                ws.append(row)
        except IndexError:
            continue

file2 = wb.save(f"{directory_path}report.xlsx")
# # --------------------------------------------------------------------------------------------------------------------
# #
# # Begin formatting --- IN PROGRESS ---
# # Unwanted rows are already gone, so lets just delete the columns we know we aren't going to use:
# # We will be using openpyxl
# #
# # First, let's set the variables for our sheet
# # Note:
# #   1) the variables bellow will eventually be removed, as we will be using existing variables from above.
xfile = openpyxl.load_workbook(filename="C:/Users/danny/Downloads/report.xlsx")
# xfile.sheetnames
sheet = xfile.active


# This function is for testing purposes only and can be removed when the program is finished. It just prints out the
#   cell values in the terminal
def print_rows():
    for row in sheet.iter_rows(values_only=True):
        print(row)


#
# Here we will delete our unwanted columns
sheet.delete_cols(idx=4, amount=3)
sheet.delete_cols(idx=5, amount=5)
# # I don't remember what this does :(
# # sheet._cells_by_row()
# #
# # # This is going to auto-set the column width - GREAT SUCCESS!!!
for col in sheet.columns:
    setlen = 0
    column = col[0].column_letter
    for cell in col:
        if len(str(cell.value)) > setlen:
            setlen = len(str(cell.value))
    set_col_width = setlen + 5
#
# sheet.column_dimensions[column].width = set_col_width
#
# #
# #
# # # This makes a whole row bold and centered - GREAT SUCCESS!!!
for cell in sheet["1:1"]:
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="center")
print_rows()

xfile.save(f"{directory_path}report_6.xlsx")

os.system("start EXCEL.EXE C:/Users/danny/Downloads/report.xlsx")

#  row = sheet["1"] **THIS REFERENCES AN ENTIRE ROW**
# sheet.delete_rows(idx = 1, amount = 7) **I WILL NEED THIS**
# sheet.delete_cols()
# Remember that rows/columns will be inserted BEFORE the given index number
# Remember that rows/columns will be deleted STARTING AT the given index number
