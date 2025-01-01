# This document is current as of 5/6/24

import time, os, csv, openpyxl
import selenium.common.exceptions
from openpyxl.styles import Font, Alignment
from selenium import webdriver
from selenium.webdriver.chrome.options import Options

course_name = []
master_report = "master.xlsx"
chrome_options = Options()
# Option to not open browser in the first place (DL reports in the background):
chrome_options.add_argument("--headless")
# Option to keep browser window open after Selenium is finished (Closes browser after using Selenium, by default.
# Keep for testing):
# chrome_options.add_experimental_option("detach", True)


# Function to navigate and login to website, navigate to page(s), and click elements to download file(s).
# Also uses select_file() to generate a dictionary with key:value pairs of course_id:file to work with, later.
def login(course_id):
    driver = webdriver.Chrome(options=chrome_options)
    course_id = course_id
    user = "102026939"
    pwd = "RedLobster1"
    # I DON'T want to have to wait for 10 seconds. 2 seconds was enough at first, but the amount of time it took for
    # the pages to load started to vary, so I had to keep increasing the implicit wait time to prevent the program
    # from crashing when it tried to click on an xpath element that was not yet loaded. It sucks, but it do be like
    # that, sometimes.
    driver.implicitly_wait(10)

    # Login
    driver.get("https://app.schoox.com/login.php")
    driver.find_element("xpath", "//*[@id='main']/div/main/input[6]").send_keys(user)
    driver.find_element("xpath", "//*[@id='main']/div/main/input[7]").send_keys(pwd)
    button_1 = driver.find_element("name", "button")
    button_1.click()

    reports = {}
    for cid in course_id:
        # url = f"https://app.schoox.com/academies/panel/dashboard2/training/course.php?acadId=7592&course_id={cid}"
        # Download report
        driver.get(f"https://app.schoox.com/academies/panel/dashboard2/training/course.php?acadId=7592&course_id={cid}")
        button_2 = driver.find_element("xpath", "/html/body/div[5]/div/div[2]/div/div[2]/div[2]/div[2]/div/img")
        button_2.click()
        button_3 = driver.find_element("xpath",
                                       "/html/body/div[5]/div/div[2]/div/div[2]/div[2]/div[2]/div/div/p[1]/a/span/span")
        button_3.click()
        # This is necessary because in the xpath; the value of the first div tag changed between div[7] to div[11],
        # one day, so if one doesn't work, we'll try the other.
        try:
            button_4 = driver.find_element("xpath",
                                           "/html/body/div[11]/div[1]/div[2]/div[2]/div[1]/generate-report/div/div/p/a")
            button_4.click()
        except selenium.common.exceptions.NoSuchElementException:
            button_4 = driver.find_element("xpath",
                                           "/html/body/div[7]/div[1]/div[2]/div[2]/div[1]/generate-report/div/div/p/a")
            button_4.click()
        # We don't want to capture the file name IMMEDIATELY because at first, it's a .tmp file, so that won't work
        time.sleep(1)
        file = select_file()
        reports[cid] = file
    return reports


# Select most recently downloaded file and return it's filepath so that it can be saved to the 'reports' dictionary
# in login()
def select_file():
    directory_path = "C:/Users/danny/Downloads/"
    most_recent_file = None
    most_recent_time = 0

    for entry in os.scandir(directory_path):
        if entry.is_file():
            mod_time = entry.stat().st_mtime_ns
            if mod_time > most_recent_time:
                most_recent_file = entry.name
                most_recent_time = mod_time
    file = f"{directory_path}{most_recent_file}"
    return file


def make_it_rain(course_id, master_report, reports):
    for num in course_id:
        course_name.append(courses[num])

    # Check for existence of master file and proceed accordingly:
    if os.path.exists(master_report):
        # If it exists: open it and check to see if a sheet exists for the selected course name.
        master = openpyxl.load_workbook(master_report)
        for cid in course_id:
            title = courses[cid]
            # If sheet exists: select sheet, transfer data from downloaded file, and format
            if title in master.sheetnames:
                master.active = master[title]
                ws = master.active
                report = reports[cid]
                convert(report, ws)
                fix(ws)
            # If not: create the sheet, transfer data from downloaded file, and format
            else:
                master.create_sheet(title)
                master.active = master[title]
                ws = master.active
                report = reports[cid]
                # print(reports)
                convert(report, ws)
                fix(ws)
    # If Master File does not exist: Create a new workbook and sheets, transfer data, and format
    else:
        master = openpyxl.Workbook()

        for cid in course_id:
            title = courses[cid]
            master.create_sheet(title)
            master.active = master[title]
            ws = master.active
            report = reports[cid]
            convert(report, ws)
            fix(ws)

        main_sheet = master['Sheet']
        main_sheet.title = 'Main Page'

    # Save our work and close the file
    save_path = f"C:/Users/danny/Project/{master_report}"
    master.save(save_path)
    master.close()

    # Open and view the master file that we just finished working with.
    os.system(f"start EXCEL.EXE {save_path}")


# This function simply converts the .csv file to an .xlsx file, and writes only the desired rows
def convert(report, ws):
    with open(report) as f:
        reader = csv.reader(f)
        for row in reader:
            try:
                if row[6] == "0" or row[6] == "% Completed":
                    # Change from making it append to making it overwrite - NOT DONE
                    ws.append(row)
            except UnicodeDecodeError:
                pass
            except IndexError:
                continue
        f.close()
        # os.remove(report) # # RESTORE AFTER TESTING!!


# This function formats the Excel workbook the way we want it. (ADD BIT TO PUT COURSE NAME AT TOP. MERGE/CENTER/BOLD)
def fix(ws):
    # Delete unwanted columns
    ws.delete_cols(idx=4, amount=3)
    ws.delete_cols(idx=5, amount=5)

    # Auto-width columns
    for col in ws.columns:
        setlen = 0
        column = col[0].column_letter
        for cell in col:
            if len(str(cell)) > setlen:
                setlen = len(str(cell))
                ws.column_dimensions[column].width = setlen

    # Bold and center the first row
    for cell in ws["1:1"]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")


# Dictionary to associate course_id with course name
courses = {
    "3846202": "NYS Intro",
    "3847508": "Anti-Harassment 1",
    "3846909": "Anti-Harassment 5",
    "3846758": "Anti-Harassment 6",
    "3848048": "Understanding Harassment 1",
    "3848322": "Understanding Harassment 2",
    "3848379": "Understanding Harassment 3",
    "3847029": "Understanding Harassment 4",
    "3846770": "Understanding Harassment 5",
    "3848626": "Understanding Harassment 6",
    "3848445": "Understanding Harassment 7",
    "4775856": "NYS Scenarios",
    "4673600": "NYC Intro",
    "4673603": "NYC Scenarios",
    "7437580": "Menu Update - FOH",
    "7437443": "Menu Update - BOH",
    "653818": "Food Safety Basics",
    "428369": "Basics of Credit Card Security",
    "597103": "Responsible Alcohol Service",
}

cids = list(courses.keys())
titles = list(courses.values())
